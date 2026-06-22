from __future__ import annotations
import io
import re
import difflib
from datetime import date, datetime
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Row positions for metadata (1-indexed)
METADATA_ROWS = {
    1: "program_name",
    2: "cohort",
    3: "instructor",
    4: "topic",
    5: "session_date",
    6: "session_time",
    7: "batch_strength",
    8: "zoom_joined",
    9: "total_responses",
}

RESPONSE_START_ROW = 14
CONFIDENCE_THRESHOLD = 0.4


def _normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9 ]", "", str(text).lower()).strip()


def _match_program(program_name: str, programs: list[dict]) -> tuple[str | None, float]:
    if not programs:
        return None, 0.0

    norm_name = _normalize(program_name)
    best_id = None
    best_score = 0.0

    for prog in programs:
        for candidate_text in (prog.get("short_name", ""), prog.get("full_name", "")):
            candidate = _normalize(candidate_text)
            seq_score = difflib.SequenceMatcher(None, norm_name, candidate).ratio()
            name_words = set(norm_name.split())
            cand_words = set(candidate.split())
            overlap = len(name_words & cand_words) / max(len(name_words | cand_words), 1)
            score = max(seq_score, overlap)
            if score > best_score:
                best_score = score
                best_id = prog["id"]

    return best_id, best_score


def parse_session_excel(file_bytes: bytes, programs: list[dict]) -> dict:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except (InvalidFileException, Exception) as e:
        raise ValueError(f"File format not recognized. Expected Zoom session export format. ({e})")

    ws = wb["Survey Report"] if "Survey Report" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows or rows[0][0] != "Program Name":
        raise ValueError("File format not recognized. Expected Zoom session export format.")

    # Extract metadata by row position
    meta: dict = {}
    for row_idx, field_name in METADATA_ROWS.items():
        row = rows[row_idx - 1]
        meta[field_name] = row[1] if len(row) > 1 else None

    # Normalize session_date to ISO string
    sd = meta.get("session_date")
    if isinstance(sd, datetime):
        meta["session_date"] = sd.date().isoformat()
    elif isinstance(sd, date):
        meta["session_date"] = sd.isoformat()
    else:
        meta["session_date"] = str(sd) if sd else None

    # Parse response rows (row 14 onwards, skip blanks)
    responses: list[dict] = []
    for row in rows[RESPONSE_START_ROW - 1:]:
        if not any(row):
            continue
        padded = (list(row) + [None] * 5)[:5]
        name, email, rating, reason, remarks = padded
        if name is None and rating is None:
            continue
        try:
            rating_int = int(rating) if rating is not None else None
        except (ValueError, TypeError):
            rating_int = None
        responses.append({
            "respondent_name": str(name).strip() if name else None,
            "respondent_email": str(email).strip() if email else None,
            "rating": rating_int,
            "reason": str(reason).strip() if reason else None,
            "remarks": str(remarks).strip() if remarks else None,
        })

    # Compute avg_rating from actual data (ignore formula cell)
    valid_ratings = [r["rating"] for r in responses if r["rating"] is not None]
    avg_rating = round(sum(valid_ratings) / len(valid_ratings), 2) if valid_ratings else None

    program_name_raw = str(meta.get("program_name") or "")
    matched_id, confidence = _match_program(program_name_raw, programs)

    return {
        "metadata": {
            "program_name": program_name_raw,
            "cohort": meta.get("cohort"),
            "instructor": str(meta["instructor"]).strip() if meta.get("instructor") else None,
            "topic": str(meta["topic"]).strip() if meta.get("topic") else None,
            "session_date": meta.get("session_date"),
            "session_time": str(meta["session_time"]).strip() if meta.get("session_time") else None,
            "batch_strength": int(meta["batch_strength"]) if meta.get("batch_strength") else None,
            "zoom_joined": int(meta["zoom_joined"]) if meta.get("zoom_joined") else None,
            "total_responses": len(responses),
            "avg_rating": avg_rating,
        },
        "responses": responses,
        "matched_program_id": matched_id,
        "match_confidence": confidence,
    }
